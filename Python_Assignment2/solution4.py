import openpyxl
from faker import Faker
import random
from openpyxl import Workbook

fake = Faker()

def generate_fake_data(num_records):
    data = []
    for _ in range(num_records):
        emp_id = fake.random_int(min=1000, max=9999)
        emp_name = fake.name()
        emp_email = fake.email()
        business_unit = random.choice(["Sales", "Marketing", "IT"])
        salary = fake.random_int(min=30000, max=150000)
        data.append((emp_id, emp_name, emp_email, business_unit, salary))
    return data

def write_to_excel(data, file_name):
    wb = Workbook()
    ws = wb.active
    ws.append(["EMP ID", "EMP NAME", "EMP EMAIL", "Business Unit", "Salary"])
    for row in data:
        ws.append(row)
    wb.save(file_name)

def read_from_excel(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data

def top_salary_employee(data):
    return max(data, key=lambda x: x[4])[1]

def top_aggregated_salary_unit(data):
    unit_salaries = {}
    for _, _, _, unit, salary in data:
        unit_salaries[unit] = unit_salaries.get(unit, 0) + salary
    return max(unit_salaries, key=unit_salaries.get)

def top_salary_employee_unit(data):
    unit_employees = {}
    for _, emp_name, _, unit, salary in data:
        if unit not in unit_employees or salary > unit_employees[unit][1]:
            unit_employees[unit] = (emp_name, salary)
    return unit_employees

def delete_least_salary_record(data):
    min_salary = min(data, key=lambda x: x[4])[4]
    data = [record for record in data if record[4] != min_salary]
    return data

def update_salary_details(data, emp_name, new_salary):
    for i, record in enumerate(data):
        if record[1] == emp_name:
            data[i] = (record[0], record[1], record[2], record[3], new_salary)
            break
    return data

# Generate fake data and write it to an Excel file
fake_data = generate_fake_data(100)
write_to_excel(fake_data, "Employee_Personal_Details.xlsx")

# Read data from the Excel file
excel_data = read_from_excel("Employee_Personal_Details.xlsx")

# Find employee with top salary
print("Employee with top salary:", top_salary_employee(excel_data))

# Find business unit with top aggregated salary
print("Business Unit with top aggregated salary:", top_aggregated_salary_unit(excel_data))

# Find employees in each Business Unit with top salary
print("Employees in each Business Unit with top salary:\n", top_salary_employee_unit(excel_data))

# Delete employee with least salary
excel_data = delete_least_salary_record(excel_data)
print("Data after deleting employee with least salary:", excel_data)

# Update John Doe's salary to $200,000 and save updated data to a new Excel file
excel_data_updated = update_salary_details(excel_data, 'John Doe', 200000)
write_to_excel(excel_data_updated, "Updated_Employee_Personal_Details.xlsx")